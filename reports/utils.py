from decimal import Decimal
from django.db.models import Sum, Count, Avg, Q, F
from datetime import datetime, timedelta
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)


def generate_report_data(report_type, filters):
    """
    Central function to generate report data based on type and filters.
    This can be used by both the API views and background tasks.
    """
    try:
        if report_type == 'supplier':
            return generate_supplier_report(filters)
        elif report_type == 'trade':
            return generate_trade_report(filters)
        elif report_type == 'invoice':
            return generate_invoice_report(filters)
        elif report_type == 'payment':
            return generate_payment_report(filters)
        elif report_type == 'depositor':
            return generate_depositor_report(filters)
        elif report_type == 'voucher':
            return generate_voucher_report(filters)
        elif report_type == 'inventory':
            return generate_inventory_report(filters)
        elif report_type == 'investor':
            return generate_investor_report(filters)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
    except Exception as e:
        logger.error(f"Error generating {report_type} report: {str(e)}", exc_info=True)
        raise


def generate_supplier_report(filters):
    """Generate supplier report data"""
    from trade.models import Trade
    
    try:
        trades = Trade.objects.filter(
            status__in=['delivered', 'completed']
        ).select_related('supplier', 'grain_type', 'buyer', 'hub')
        
        # Apply filters
        trades = apply_date_filters(trades, filters)
        trades = apply_hub_filter(trades, filters)
        
        if filters.get('supplier_id'):
            trades = trades.filter(supplier_id=filters['supplier_id'])
        if filters.get('grain_type_id'):
            trades = trades.filter(grain_type_id=filters['grain_type_id'])
        
        # Aggregate by supplier
        supplier_data = trades.values(
            'supplier_id',
            'supplier__first_name',
            'supplier__last_name',
            'supplier__phone_number'
        ).annotate(
            total_trades=Count('id'),
            total_quantity_kg=Sum('quantity_kg'),
            total_value=Sum('total_trade_cost'),
            avg_price_per_kg=Avg('buying_price')
        ).order_by('-total_quantity_kg')
        
        if filters.get('min_total_supplied'):
            supplier_data = supplier_data.filter(
                total_quantity_kg__gte=filters['min_total_supplied']
            )
        
        return list(supplier_data)
    except Exception as e:
        logger.error(f"Error in generate_supplier_report: {str(e)}", exc_info=True)
        raise



def generate_payment_report(filters):
    """Generate payment report data - FIXED"""
    from accounting.models import Payment
    
    try:
        payments = Payment.objects.select_related(
            'invoice', 'invoice__account', 'created_by'
        ).filter(status='completed')
        
        # Apply filters
        if filters.get('start_date'):
            payments = payments.filter(payment_date__gte=filters['start_date'])
        if filters.get('end_date'):
            payments = payments.filter(payment_date__lte=filters['end_date'])
        
        # ✅ FIX: Properly handle payment_method filter (can be empty list)
        payment_method_filter = filters.get('payment_method', [])
        if payment_method_filter and isinstance(payment_method_filter, list) and len(payment_method_filter) > 0:
            payments = payments.filter(payment_method__in=payment_method_filter)
        
        if filters.get('account_id'):
            payments = payments.filter(invoice__account_id=filters['account_id'])
        if filters.get('min_amount'):
            payments = payments.filter(amount__gte=filters['min_amount'])
        
        return list(payments)
    except Exception as e:
        logger.error(f"Error in generate_payment_report: {str(e)}", exc_info=True)
        raise



def generate_voucher_report(filters):
    """Generate voucher report data - FIXED"""
    from vouchers.models import Voucher
    
    try:
        vouchers = Voucher.objects.select_related(
            'deposit', 'deposit__farmer', 'deposit__hub',
            'deposit__grain_type', 'holder'
        )
        
        # Apply filters
        if filters.get('start_date'):
            vouchers = vouchers.filter(issue_date__gte=filters['start_date'])
        if filters.get('end_date'):
            vouchers = vouchers.filter(issue_date__lte=filters['end_date'])
        if filters.get('hub_id'):
            vouchers = vouchers.filter(deposit__hub_id=filters['hub_id'])
        
        # ✅ FIX: Properly handle status filters (can be empty lists)
        status_filter = filters.get('status', [])
        if status_filter and isinstance(status_filter, list) and len(status_filter) > 0:
            vouchers = vouchers.filter(status__in=status_filter)
        
        verification_status_filter = filters.get('verification_status', [])
        if verification_status_filter and isinstance(verification_status_filter, list) and len(verification_status_filter) > 0:
            vouchers = vouchers.filter(verification_status__in=verification_status_filter)
        
        if filters.get('holder_id'):
            vouchers = vouchers.filter(holder_id=filters['holder_id'])
        if filters.get('grain_type_id'):
            vouchers = vouchers.filter(deposit__grain_type_id=filters['grain_type_id'])
        
        return list(vouchers)
    except Exception as e:
        logger.error(f"Error in generate_voucher_report: {str(e)}", exc_info=True)
        raise



def generate_investor_report(filters):
    """Generate investor report data"""
    from investors.models import InvestorAccount
    
    try:
        accounts = InvestorAccount.objects.select_related('investor').prefetch_related(
            'trade_financings', 'trade_loans', 'deposits', 'withdrawals'
        )
        
        # Apply filters
        if filters.get('investor_id'):
            accounts = accounts.filter(investor_id=filters['investor_id'])
        if filters.get('min_total_invested'):
            accounts = accounts.filter(total_utilized__gte=filters['min_total_invested'])
        
        return list(accounts)
    except Exception as e:
        logger.error(f"Error in generate_investor_report: {str(e)}", exc_info=True)
        raise


def generate_trade_report(filters):
    """Generate trade report data - FIXED VERSION"""
    from trade.models import Trade
    
    try:
        # Select related to avoid N+1 queries
        trades = Trade.objects.select_related(
            'buyer', 'supplier', 'grain_type', 'quality_grade', 
            'hub', 'initiated_by'
        ).prefetch_related('grns')
        
        # Apply date filters on created_at
        if filters.get('start_date'):
            trades = trades.filter(created_at__date__gte=filters['start_date'])
        if filters.get('end_date'):
            trades = trades.filter(created_at__date__lte=filters['end_date'])
        
        # Apply hub filter
        if filters.get('hub_id'):
            trades = trades.filter(hub_id=filters['hub_id'])
        
        # Apply status filter - handle empty list properly
        # status_filter = filters.get('status', [])
        # if status_filter and isinstance(status_filter, list) and len(status_filter) > 0:
        #     trades = trades.filter(status__in=status_filter)

        status_filter = filters.get('status', [])
        if status_filter and len(status_filter) > 0:
            trades = trades.filter(status__in=status_filter)
        
        # Apply other filters
        if filters.get('buyer_id'):
            trades = trades.filter(buyer_id=filters['buyer_id'])
        if filters.get('supplier_id'):
            trades = trades.filter(supplier_id=filters['supplier_id'])
        if filters.get('grain_type_id'):
            trades = trades.filter(grain_type_id=filters['grain_type_id'])
        
        # Value filters
        if filters.get('min_value'):
            trades = trades.filter(total_trade_cost__gte=filters['min_value'])
        if filters.get('max_value'):
            trades = trades.filter(total_trade_cost__lte=filters['max_value'])
        
        # Order by creation date
        trades = trades.order_by('-created_at')
        
        return list(trades)
        
    except Exception as e:
        logger.error(f"Error in generate_trade_report: {str(e)}", exc_info=True)
        raise


# def generate_invoice_report(filters):
#     from accounting.models import Invoice
#     try:
#         invoices = Invoice.objects.select_related('account', 'trade', 'trade__buyer')

#         invoices = apply_date_filters(invoices, filters, 'issue_date')
#         invoices = apply_hub_filter(invoices, filters, 'trade__hub')

#         if filters.get('account_id'):
#             invoices = invoices.filter(account_id=filters['account_id'])

#         if filters.get('payment_status'):
#             invoices = invoices.filter(payment_status__in=filters['payment_status'])

#         # FIXED: proper boolean check
#         if filters.get('overdue_only') is True:
#             invoices = invoices.filter(
#                 payment_status='overdue',
#                 due_date__lt=timezone.now().date()
#             )

#         if filters.get('min_amount'):
#             invoices = invoices.filter(total_amount__gte=filters['min_amount'])

#         return list(invoices)
#     except Exception as e:
#         logger.error(f"Error in generate_invoice_report: {str(e)}", exc_info=True)
#         raise

def generate_invoice_report(filters):
    """✅ FIXED: Properly handle empty arrays"""
    from accounting.models import Invoice
    
    try:
        invoices = Invoice.objects.select_related('account', 'trade', 'trade__buyer')
        
        # Apply date filters
        if filters.get('start_date'):
            invoices = invoices.filter(issue_date__date__gte=filters['start_date'])
        if filters.get('end_date'):
            invoices = invoices.filter(issue_date__date__lte=filters['end_date'])
        
        # Apply hub filter
        if filters.get('hub_id'):
            invoices = invoices.filter(trade__hub_id=filters['hub_id'])
        
        if filters.get('account_id'):
            invoices = invoices.filter(account_id=filters['account_id'])
        
        # ✅ FIX: Properly handle payment_status filter
        payment_status_filter = filters.get('payment_status', [])
        if payment_status_filter and len(payment_status_filter) > 0:
            invoices = invoices.filter(payment_status__in=payment_status_filter)
        
        # Proper boolean check
        if filters.get('overdue_only') is True:
            invoices = invoices.filter(
                payment_status='overdue',
                due_date__lt=timezone.now().date()
            )
        
        if filters.get('min_amount'):
            invoices = invoices.filter(total_amount__gte=filters['min_amount'])
        
        return list(invoices)
        
    except Exception as e:
        logger.error(f"Error in generate_invoice_report: {str(e)}", exc_info=True)
        raise


def generate_depositor_report(filters):
    from vouchers.models import Deposit
    try:
        deposits = Deposit.objects.select_related('farmer', 'grain_type', 'hub')

        deposits = apply_date_filters(deposits, filters, 'deposit_date')
        deposits = apply_hub_filter(deposits, filters)

        if filters.get('farmer_id'):
            deposits = deposits.filter(farmer_id=filters['farmer_id'])
        if filters.get('grain_type_id'):
            deposits = deposits.filter(grain_type_id=filters['grain_type_id'])

        # FIXED
        if filters.get('validated_only') is True:
            deposits = deposits.filter(validated=True)

        if filters.get('min_total_quantity'):
            deposits = deposits.annotate(total_qty=Sum('quantity_kg')) \
                           .filter(total_qty__gte=filters['min_total_quantity'])

        return list(deposits)
    except Exception as e:
        logger.error(f"Error in generate_depositor_report: {str(e)}", exc_info=True)
        raise


def generate_inventory_report(filters):
    from vouchers.models import Inventory
    try:
        inventories = Inventory.objects.select_related('grain_type', 'hub')

        inventories = apply_hub_filter(inventories, filters)

        if filters.get('grain_type_id'):
            inventories = inventories.filter(grain_type_id=filters['grain_type_id'])
        if filters.get('min_quantity'):
            inventories = inventories.filter(total_quantity_kg__gte=filters['min_quantity'])

        # FIXED
        if filters.get('low_stock_only') is True:
            inventories = inventories.filter(
                available_quantity_kg__lt=F('total_quantity_kg') * Decimal('0.2')
            )

        return list(inventories)
    except Exception as e:
        logger.error(f"Error in generate_inventory_report: {str(e)}", exc_info=True)
        raise


def apply_date_filters(queryset, filters, date_field='created_at'):
    if filters.get('start_date'):
        queryset = queryset.filter(**{f'{date_field}__date__gte': filters['start_date']})
    if filters.get('end_date'):
        queryset = queryset.filter(**{f'{date_field}__date__lte': filters['end_date']})
    return queryset


def apply_hub_filter(queryset, filters, hub_field='hub'):
    if filters.get('hub_id'):
        queryset = queryset.filter(**{hub_field: filters['hub_id']})
    return queryset


def calculate_aging(invoices):
    """Calculate accounts receivable aging"""
    aging = {
        'current': Decimal('0.00'),
        '1-30_days': Decimal('0.00'),
        '31-60_days': Decimal('0.00'),
        '61-90_days': Decimal('0.00'),
        'over_90_days': Decimal('0.00')
    }
    
    today = timezone.now().date()
    
    for invoice in invoices:
        if invoice.payment_status == 'paid':
            continue
        
        days_overdue = (today - invoice.due_date).days if today > invoice.due_date else 0
        amount = invoice.amount_due
        
        if days_overdue <= 0:
            aging['current'] += amount
        elif days_overdue <= 30:
            aging['1-30_days'] += amount
        elif days_overdue <= 60:
            aging['31-60_days'] += amount
        elif days_overdue <= 90:
            aging['61-90_days'] += amount
        else:
            aging['over_90_days'] += amount
    
    return aging


def export_to_csv(data, columns):
    """Export data to CSV format"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    
    for row in data:
        writer.writerow({col: row.get(col, '') for col in columns})
    
    return output.getvalue()


def export_to_excel(data, columns, sheet_name='Report'):
    """Export data to Excel format"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, column in enumerate(columns, 1):
                ws.cell(row=row_num, column=col_num).value = row_data.get(column, '')
        
        # Auto-adjust column widths
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


def export_to_pdf(data, columns, title='Report'):
    """Export data to PDF format"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Table data
        table_data = [columns]
        for row in data:
            table_data.append([str(row.get(col, '')) for col in columns])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        return output.getvalue()
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")


def format_currency(amount):
    """Format amount as currency"""
    return f"UGX {amount:,.2f}"


def format_percentage(value):
    """Format value as percentage"""
    return f"{value:.2f}%"


def format_date(date_obj):
    """Format date consistently"""
    if isinstance(date_obj, str):
        return date_obj
    return date_obj.strftime('%Y-%m-%d') if date_obj else ''


def format_datetime(datetime_obj):
    """Format datetime consistently"""
    if isinstance(datetime_obj, str):
        return datetime_obj
    return datetime_obj.strftime('%Y-%m-%d %H:%M:%S') if datetime_obj else ''